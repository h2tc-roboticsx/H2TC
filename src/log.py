import os, sys, csv, time, shutil
from openpyxl import load_workbook, Workbook
from addict import Dict
from itertools import product
from random import choices, choice, seed, randint


'''
Path

'''

# # path to the current source code file
# curr_path = "/media/ur-5/golden_t/data/throw"
# # path to the data directory
# data_path = os.path.join(curr_path, "data")
# # create data directory if not exists
# if not os.path.isdir(data_path):
#     os.mkdir(data_path)
    
# path to the code repository root
root_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..")
# path to the registry directory with objects and subjects info
regs_path = os.path.join(root_path, "register")

# path to the logbook
log_path = os.path.join(root_path, "log.xlsx")
# path to backup the logbook
log_backup_path = os.path.join(root_path, "log_backup.xlsx")


'''
Constants

'''

SUBJECTS = [] # list of all subjects' ID
with open(os.path.join(regs_path, "subjects.csv")) as f:
    # load subject IDs from the file 'register/subjects.csv'
    for line in csv.reader(f, delimiter='\n'):
        # each line represents a subject
        SUBJECTS.append(line[0])

OBJECTS = [] # list of all objects' ID
with open(os.path.join(regs_path, 'objects.csv')) as f:
    # load object IDs from the file 'register/objects.csv'
    lines = csv.reader(f)
    for line in lines: # iterate every line in the file
        # separate the line content by '\t'
        line = line[0].split('\t')
        if line[-1] == '0':
            # objects without marker
            OBJECTS.append("{} ({})".format(line[0], line[1]))
        else:
            # objects with marker
            OBJECTS.append("{} ({}) marked".format(*line))

# options of hand vertical position
HEIGHT = ['overhead', 'overhand', 'chest', 'underhand']
# options of throwing speed
SPEED = ['fast', 'normal', 'slow']
# options of hand horizontal position
HORIZON = ['left', 'middle', 'right']
# options of subject action
ACTION = ['throw', 'catch']
# options of hand used to throw/catch
HAND = ['single', 'both']
# options of either wearing equipment or not
EQUIPPED = [1, 0]

# headers in the spreadsheet
HEADERS = ['no', 'object', 'equipped', 'action', 'hand', 'position', 'height',
           'horizon', 'speed', 'take_id', 'success', 'verified', 'annotated']

# coordinates size for positioning the subject in the zone
# 4x4: width and height are divided into 4 units respectively
GROUND_SIZE = (4, 4)

# quick conversion from English to Chinese
CHN = {
    None : '',
    '' : '',
    'overhead' : '头顶',
    'overhand' : '肩顶',
    'chest' : '胸前',
    'underhand' : '腹部',
    'fast' : '快速',
    'slow' : '慢速',
    'normal' : '正常速度',
    'left' : '左侧',
    'right' : '右侧',
    'middle' : '中间',
    'throw' : '抛',
    'catch' : '接',
    'single' : '单手',
    'both' : '双手'
}

# default string for the missing value
EMPTY = ''


'''
Help functions

'''

def init_subject_sheet(ws):
    '''
    Initialize the content of a subject sheet

    Returns:
        None

    Args:
        ws (openpyxl.Worksheet): the sheet to be written 
    
    '''

    # iterate every header in the predefined HEADER collection
    # index starts from 1 as column number in the sheet
    for col, header in enumerate(HEADERS, 1):
        # write header string at the first row, 'col' column
        ws.cell(column=col, row=1).value = header

    i = 2 # row index starting from 2
    # iterate all instructions: OBJECTS x Equipped or Not x Throw or Catch
    for obj, equipped, action in product(OBJECTS, EQUIPPED, ACTION):
        if equipped == 1: # the subject wearing equipment
            # hand: free, free, free, random, random
            hands = [EMPTY, EMPTY, EMPTY] + choices(HAND, k=2)
        else: # the subject not wearing equipment
            # hand: free, single, both, single, both
            hands = [EMPTY] + HAND + HAND

        # iterate every hand status
        for hand in hands:
            # set take info at the row i in the sheet
            set_value(ws,
                      i, # row
                      no=i-1, # take numbering in the sheet
                      object=obj, # object ID
                      equipped=equipped, # either equipped or not
                      hand=hand, # hand status
                      action=action)
            i += 1 # increase the row by 1
            
def value(ws, header, row, default=EMPTY, headers=HEADERS):
    '''
    Get the value of the header at the row in the sheet

    Returns:
        auto: the value

    Args:
        ws (openpyxl.Worksheet): sheet to be accessed
        header (string): the header of the column to be accessed
        row (int): the row number to be accessed
        default (auto): the default value to be returned if not found

    '''

    # get the value of the header at the row
    # column index is retrieved by the index of the header in HEADERS
    v = ws.cell(column=headers.index(header)+1, row=row).value
    # return default value if the value returned from the sheet is None
    # otherwise, return the original value
    return default if v is None else v
    
def set_value(ws, row, headers=HEADERS, **kwargs):
    '''
    Set values at the row in the sheet

    Returns:
        None

    Args:
        ws (openpyxl.Worksheet): the sheet to be written
        row (int): the row to be written at
        kwargs (dict): key is the header, value to be written

    '''

    for k, v in kwargs.items(): # iterate each key and value in kwargs
        # use the key as the header
        # set the value at the row under the header in the sheet
        ws.cell(column=headers.index(k)+1, row=row).value = v

def finished(ws, row):
    '''
    Check if the recording of the take at the row is finished

    Returns:
        bool: true for finished
    
    Args:
        ws (openpyxl.Worksheet): the sheet to be checked
        row (int): the row where the take is stored

    '''

    # true if the take ID is not EMPTY
    return value(ws, 'take_id', row) != EMPTY

def annotated(ws, row):
    '''
    Check if the take at the row is annotated

    Returns:
        bool: true for annotated
    
    Args:
        ws (openpyxl.Worksheet): the sheet to be checked
        row (int): the row where the take is stored

    '''

    # true if the value of 'annotated' is not EMPTY
    return value(ws, 'annotated', row) != EMPTY

def problematic(ws, row):
    '''
    Check if the take at the row is problematic

    Returns:
        bool: true for having issues
    
    Args:
        ws (openpyxl.Worksheet): the sheet to be checked
        row (int): the row where the take is stored

    '''

    # true if the 'verified' value is -1
    return value(ws, 'verified', row) in ['-1', -1]

def failed(ws, row):
    '''
    Check if the take at the row is failed

    Returns:
        bool: true for failed
    
    Args:
        ws (openpyxl.Worksheet): the sheet to be checked
        row (int): the row where the take is stored

    '''

    # true if the 'success' value is 0
    return value(ws, 'success', row) in ['0', 0]

def save(log, path, backup_path=log_backup_path):
    '''
    Save the logbook

    Returns:
        None

    Args:
        log (openpyxl.Workbook): logbook to save
        path (string): path to save
        backup_path (string): path to backup the logbook

    '''
    
    if os.path.isfile(path):
        # copy the file to the backup path as a backup
        shutil.copy(path, backup_path)

    # save the logbook
    log.save(path)

def take_id_to_str(tid):
    '''
    Convert take ID from int or incomplete string to string

    Full take ID string: e.g. 100 to 000100

    Returns:
        string: string of full take ID

    Args:
        tid (int/string): take ID int / incomplete string

    '''
    
    if isinstance(tid, int):
        return '{:06d}'.format(tid)
    else:
        return '{:06d}'.format(int(tid))


def parse_convert_take_ids(take_ids):
    '''
    parse and complement takd IDs

    Return:
        list: full take ID string

    Args:
        take_ids (list): list of take id abbreviations 

    '''

    # list of take IDs to be returned
    takes = []
    # iterate through every take ID abbreviation
    for take_id in take_ids:
        if '-' in take_id:
            # complement e.g. '10-12' to a list of 10, 11, 12
            # if '-' detected in take ID abbreviation
            take_id = take_id.split('-')
            assert len(take_id) == 2
            takes += list(range(int(take_id[0]), int(take_id[1])+1))
        else:
            takes.append(take_id)
    
    return [take_id_to_str(take_id) for take_id in takes]


'''
Log class

'''
    
class Log:
    '''
    Encapsule the methods for accessing and modifying log data.
    The internal data structure of log is an Excel spreadsheet

    It also generates the instructions for the subjects.

    '''
    
    def __init__(self, path=log_path):
        '''
        Initialization of the class

        Returns:
            None
        
        Args:
            path (string): path to the logbook file

        '''
        
        self.path = path
        # load the logbook file if exists, otherwise, create a new empty one
        self.logbook = load_workbook(path) if os.path.isfile(path) else Workbook()
        # filter out sheets not belonging to subjects
        for sub in self.logbook.sheetnames:
            if sub not in SUBJECTS:
                # remove the sheet whose name is not a valid subject ID in SUBJECTS
                self.logbook.remove(self.logbook[sub])

        # get all folders (name) in the data directory
        # sort by the folder name as integer
        fnames = sorted(os.listdir(data_path), key=int)
        # take ID = 0 if no folder in the data directory
        # otherwise, take ID = the last take (folder) ID + 1
        self.take_id = 0 if len(fnames) == 0 else int(fnames[-1])+1

    def sheet(self, subject):
        '''
        Get the sheet by the subject name

        Returns:
            openpyxl.Worksheet: spread sheet instance of the subject

        Args:
            subject (string): subject ID

        '''
        
        if subject not in self.logbook.sheetnames: # no sheet named the subject ID
            # create a new sheet
            ws = self.logbook.create_sheet(subject)
            # initialize sheet content
            init_subject_sheet(ws)
            return ws
        else: # having sheet named the subject ID
            return self.logbook[subject]

    def __unfinished(self, subjects):
        '''
        Find available and unfinished tasks with instructions for a pair of subjects

        Returns:
            iterator: iterator with unfinished takes

        Args:
            subjects (tuple): a typle of two strings of subject IDs

        '''
        
        sub1, sub2 = subjects
        # get sheets for the subjects
        ws1, ws2 = [self.sheet(subject) for subject in subjects]

        takes = [] # list of the take dict with instruction content
        
        for row in range(2, ws1.max_row): # iterate every row in subject1's sheet from 2
            # skip the row (take) if it has been finished for subject 1
            if finished(ws1, row):
                continue
            
            # skip the takes where subject1 is not equipped
            if value(ws1, 'equipped', row) == 0:
                continue

            # we want the row (take): not finished and the subject 1 wearing the equipment
            
            # find the paired row in the sheet of subject 2
            # the paired row = subject 1 row + 15 if subject 2 is catcher
            # the paired row = subject 1 row + 5 if subject 2 is thrower
            paired = row + 15 if value(ws1, 'action', row) == 'throw' else row + 5

            # skip if the take has been finished for subject 2
            if finished(ws2, paired):
                continue

            # change the random seed according to time to generate real random number
            seed(time.time())
            
            # take information dict
            # key: subject ID
            # value: dict of instruction and other info like row in the corresponding sheet
            take = Dict({sub1 : {'row':row}, sub2 : {'row':paired}})
            # log object info
            take.obj = value(ws1, 'object', row)

            # for each subject ID, subject sheet, and row number
            for sub, ws, i in zip(subjects, (ws1, ws2), (row, paired)):
                # randomly initialize the subject's position
                take[sub].position = str((randint(0, GROUND_SIZE[0]-1), randint(0, GROUND_SIZE[1]-1)))
                # store hand option in the take dict
                take[sub].hand = value(ws, 'hand', i)
                take[sub].action = value(ws, 'action', i)

            # remove instruction if both subjects are free to behave i.e. empty hand
            if take[sub1].hand == EMPTY and take[sub2].hand == EMPTY:
                take[sub2].height = EMPTY
                take[sub2].speed = EMPTY
                take[sub2].horizon = EMPTY
            else: # not free instruction
                # randomly sample hand vertical position
                take[sub2].height = choice(HEIGHT)
                if take[sub2].action == 'throw': # subject 2 throws
                    # randomly sample throwing speed
                    take[sub2].speed = choice(SPEED)
                else: # subject 2 catches
                    # randomly sample hand horizontal position
                    take[sub2].horizon = choice(HORIZON)

            # add the take dict to the list
            takes.append(take)

        # return an iterator of the take dict list
        return iter(takes)

    def next_take(self, subjects):
        '''
        Get the next take to capture

        Returns:
            dict: take info to be captured

        Args:
            subjects (tuple): a tuple of two strings of subject IDs

        '''
        
        if not hasattr(self, 'subjects'): # subjects attribute not exists
            self.subjects = subjects
            # get unfinished takes for this pair of subjects
            self.takes = self.__unfinished(subjects)

        # regenerate unfinished takes if the argument subjects are
        # different from the attribute subjects
        if any(_sub != sub for _sub, sub in zip(self.subjects, subjects)):
            self.takes = self.__unfinished(subjects)
            self.subjects = subjects
            
        try:
            take = next(self.takes) # get next take dict in the iterator
            # convert take ID from int to string
            take.ID = take_id_to_str(self.take_id)
        except Exception:
            # set take to None if any exception
            take = None
        return take

    def finish(self, rec_id):
        '''
        Update the ID for the next take and save the logbook
        this should be called as the recording of the current take finished
        
        Returns:
            None
        
        Args:
            rec_id (string): the ID of the take recorded

        '''

        # convert ID string to int
        rec_id = int(rec_id)

        # increase take ID by 1 if the current recording is the current take
        if rec_id == self.take_id:
            self.take_id += 1

        # save the logbook
        self.save()
        
    def log(self, ws, row, **kwargs):
        '''
        Write the content of kwargs at a specified row in the sheet
        
        Returns:
            None
    
        Args:
            ws (string/openpyxl.Worksheet): (name of) the sheet to be written
            row (int): which row the data should be written at
            kwargs (dict): key is the header in the sheet

        '''
        
        if isinstance(ws, str):
            # get the sheet instance if the given ws is string
            ws = self.logbook.get_sheet_by_name(ws)

        # set the values in the sheet
        set_value(ws, row, **kwargs)
    
    def save(self):
        '''
        Save the logbook with backup

        Returns:
            None

        Args:
            None

        '''

        if os.path.isfile(self.path): # the logbook file exists
            # copy the file to the backup path as a backup
            shutil.copy(self.path, log_backup_path)

        # save the logbook to the path
        self.logbook.save(self.path)
        
