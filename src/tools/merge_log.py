import os, sys
from openpyxl import load_workbook
from argparse import ArgumentParser

# get the absolute path of the current and root directory
curr_path = os.path.dirname(os.path.abspath(__file__))
root_path = os.path.join(curr_path, "../..")

# add root directory to the system path
sys.path.append(root_path)

from src.log import finished, value, set_value, save, EMPTY

argparser = ArgumentParser()
argparser.add_argument("--old", default="old.xlsx",
                       help="the name of the old logbook file")
argparser.add_argument("--new", default="log.xlsx",
                       help="the name of the new logbook file")
argparser.add_argument("-o", "--output", default="log.xlsx",
                       help="filename of the merged logbook file")

args = argparser.parse_args()

# full path to the old logbook file
old_log_path = os.path.join(root_path, args.old)
new_log_path = os.path.join(root_path, args.new)
# path to save the merged log
save_path = os.path.join(root_path, args.output)

# load logbook from the files
old_log = load_workbook(old_log_path)
new_log = load_workbook(new_log_path)

# headers to be merged
headers = ['verified', 'success', 'annotated']

for sub in old_log.sheetnames: # iterate every subject sheet
    old_sheet = old_log[sub]
    new_sheet = new_log[sub]
    
    for row in range(2, old_sheet.max_row): # iterate every row from second in the old sheet
        # skip the unfinished take since they don't have value for the specified headers
        if not finished(old_sheet, row): continue
        
        take_id = value(old_sheet, 'take_id', row)
        # value of headers to be in the output logbook
        values = {}
        for header in headers:
            old_value = value(old_sheet, header, row)
            new_value = value(new_sheet, header, row)

            # skip this header if both have the same value
            if old_value == new_value: continue

            print(take_id, header, old_value, new_value)
            
            # if header == 'success':
            #     print('Inconsistent success of {}: {} {}'.format(take_id, old_value, new_value))
            #     # set success to be 0 (failed) if inconsistent
            #     # because this currently can be only caused by someone manually setting
            #     # success to failed when the take is observed to be failed durign annotation
            #     values[header] = '0'
            # else:
            #     # set the value to the non-empty one
            #     values[header] = old_value if old_value != EMPTY else new_value
            
        # set_value(new_sheet, row, **values)
            
# save the merged logbook
# save(new_log, save_path)
