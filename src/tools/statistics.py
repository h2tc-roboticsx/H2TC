import os, sys, json
from openpyxl import load_workbook, Workbook
from addict import Dict

# get the absolute path of the current and root directory
curr_path = os.path.dirname(os.path.abspath(__file__))
root_path = os.path.join(curr_path, "../..")

# add root directory to the system path
sys.path.append(root_path)

from src.log import value, set_value, data_path, log_path, EMPTY
from src.annotate import anno_dir

# headers of the statistics data Excel file
STAT_HEADERS = ['take_id', 'success', 'annotated', 'verified', 'object', 'length',
                'left', 'right', '115', '116', '117', '118', '101', '102', '103', '104', '105', '106', '107', '108',
                '109', '110', '111', '112', '113', '114', '119']

# directory to save the statistics data
stat_path = os.path.join(root_path, 'statistics')
if not os.path.exists(stat_path):
    os.mkdir(stat_path)

# path to the raw statistics data dict
stats_raw_path = os.path.join(stat_path, 'statistics.json')
# path to the statistics data Excel file
stats_sheet_path = os.path.join(stat_path, 'statistics.xlsx')
obj_dist_path = os.path.join(stat_path, 'objects.json')


'''
Computing drop rate for HE and OptiTrack

'''

def drop_rate_optitrack(fpath, total=1200):
    '''
    calculate the drop rate of optitrack data for each object

    Returns:
        dict: key (string): object ID in optitrack system
              value (float): drop rate

    Args:
        fpath (string): path to the optitrack data file
        total (int): the virtual total amount of frames calculated by RECORDING_DURATION * FPS

    '''

    # frames counter for each object
    frames = {}
    # load optitrack data from data file
    with open(fpath, 'r') as f: lines = f.readlines()
    
    for line in lines: # iterate through each line
        # parse the first word as object ID
        obj = line.split(' ')[0]
        # count the frame 
        if obj in frames:
            frames[obj] += 1
        else:
            frames[obj] = 1

    # init the dict to store the drop rates for each object
    drops = {}
    # update the expected total amount of frames as the maximum amount of frames
    # in any object if it is greater than the virtual total amount (passed from arguments)
    for _, num_frames in frames.items(): total = max(total, num_frames)
    
    for obj, num_frames in frames.items():
        drop_rate = float(total - num_frames) / total
        drops[obj] = drop_rate
    return drops

def drop_rate_he(he_path, total=600):
    '''
    calculate the drop rate of HE data for left and right hands

    Returns:
        dict: key (string): left and right
              value (float): drop rate

    Args:
        fpath (string): path to the HE data directory
        total (int): the virtual total amount of frames calculated by RECORDING_DURATION * FPS

    '''

    drops = {} # the dict to store the drop rates for two hands
    nframes = {} # frames counter
    for hand, raw in [('left', 'P1L.csv'), ('right', 'P1R.csv')]:
        # 'raw' denotes the name of the raw data file
        fname = os.path.join(he_path, raw)
        # load raw data from the file
        with open(fname, 'r') as f: frames = f.readlines()
        # the total amount of frames = number of rows - 1 (headers)
        num_frames = len(frames)-1
        nframes[hand] = num_frames
        
        # update the expected total amount of frames as the maximum amount of frames
        # in any hand if it is greater than the virtual total amount (passed from arguments)
        total = max(total, num_frames)

    # calculate drop rate for each hand
    for hand, num_frames in nframes.items():
        drop_rate = float(total - num_frames) / total
        drops[hand] = drop_rate
    return drops


'''
Main procedure

'''

if __name__ == '__main__':
    # load logbook from the path
    log = load_workbook(log_path)
    
    if os.path.isfile(stats_raw_path): # the statistics data exists
        # load statistics data from the file
        with open(stats_raw_path, 'r') as f:
            stats = Dict(json.loads(f.read()))
    else: # the statistics data doesn't exist
        stats = Dict()

    status = Dict({
        'success' : 0,
        'annotated' : 0,
        'problem' : 0,
    })

    objects = {}

    '''
    gather data from logbook to the statistics
    
    '''
    for sheet in log.sheetnames: # iterate the sheet of every subject
        # get the sheet instance from the sheet name
        ws = log[sheet]
        for row in range(2, ws.max_row): # iterate every row in the sheet from the second row
            take_id = value(ws, 'take_id', row)
            # skip the take if it is not finished
            if take_id == EMPTY: continue

            # init the statistics data dict for the current take
            if take_id not in stats:
                stats[take_id] = Dict()
            
            for header in STAT_HEADERS[:5]: # iterate the first 5 entries in the predefined HEADERS
                v = value(ws, header, row)
                
                # convert the value to an int if the header is any of the above three
                if header in ['success', 'verified', 'annotated']:
                    v = v if v == EMPTY else int(v)
                    
                stats[take_id][header] = v

    '''
    Count the drop rate of frames for the raw data under the data path
    
    '''
    for take_id in os.listdir(data_path): # iterate every take folder in the data path
        # skip the take if its statistics is not counted
        if take_id not in stats: continue
        
        stat = stats[take_id]
        # the path to the take data
        take_path = os.path.join(data_path, take_id)

        if 'left' not in stat: # the HE frame drop is not counted
            # path to the HE data folder under the raw directory
            fpath = os.path.join(take_path, 'raw/hand')
            if os.path.exists(fpath):
                # store drop rate in the statistics dict
                for k, v in drop_rate_he(fpath).items(): stat[k] = v

            # path to the HE data folder under the take directory
            # note that only one of these two paths will really exist and have the data
            fpath = os.path.join(take_path, 'hand')
            if os.path.exists(fpath):
                for k, v in drop_rate_he(fpath).items(): stat[k] = v
                
        if '115' not in stat: # optitrack frame drop is not counted
            # path to the optitrack data file under the raw direcotry
            fpath = os.path.join(take_path, 'raw/optitrack.csv')
            if os.path.isfile(fpath):
                # store drop rate in the statistics dict
                for k, v in drop_rate_optitrack(fpath).items(): stat[k] = v

            # path to the optirack data file under the take folder
            # note that only one of these two paths will really exist and have the data
            fpath = os.path.join(take_path, 'optitrack.csv')
            if os.path.isfile(fpath):
                for k, v in drop_rate_optitrack(fpath).items(): stat[k] = v

    '''
    Get real annotation status from the annotation data
    
    '''    
    anno_status = {}
    for anno_filename in os.listdir(anno_dir):
        # full path to the annotation data file
        anno_file = os.path.join(anno_dir, anno_filename)
        # load annotation data from the file
        with open(anno_file, 'r') as f: anno = json.loads(f.read())
        # parse take ID from the file name
        take_id = anno_filename[:-5]
        anno_status[take_id] = int(anno['status'])
        
    # all take IDs in the statistics data dict
    take_ids = sorted(stats.keys(), key=int)
    # init new Excel book to visualize the data
    wb = Workbook()
    # get the default sheet in the book
    ws = wb.active
    
    # init headers in the first row of the sheet
    for col, header in enumerate(STAT_HEADERS, 1):
        ws.cell(column=col, row=1).value = header
        
    for row, take_id in enumerate(take_ids, 2): # iterate every take with index starting from 2
        stat = stats[take_id]
        # set values in the 'stat' in the sheet for the corresponding headers
        set_value(ws, row, STAT_HEADERS, **stat)

        # assert the valid values of the below headers
        assert stat.success in (0, 1)        
        assert stat.annotated in (EMPTY, None, 1)
        assert stat.verified in (-1, 1, None, EMPTY)
        
        # count the number of success
        if stat.success == 1:
            status.success += 1
            
        if stat.annotated==1 or stat.verified==-1: # take is either annotated or problematic
            # path to the annotation data file
            anno_file_path = os.path.join(anno_dir, take_id + '.json')
            # report if the annotation data file is missing
            if not os.path.exists(anno_file_path):
                print("Missing annotation data: ", take_id)
                continue
            
        # count the number of annotated
        if stat.annotated == 1:
            status.annotated += 1
            assert anno_status[take_id] == 1
            del anno_status[take_id]
            
        # count the number of problematic
        if stat.verified == -1:
            status.problem += 1
            assert anno_status[take_id] == -1
            del anno_status[take_id]
            

        # count object distribution
        obj = stat.object
        if obj in objects:
            objects[obj] += 1
        else:
            objects[obj] = 1

    for take_id, _status in anno_status.items():
        if _status == 0:
            print("Annotation data exists but not finished: {}".format(take_id))
        else:
            print(_status)
            print("Status in the annotation data is inconsistent to the log records: {}"
                  .format(take_id))
        
    # total amount of takes
    tol = len(take_ids)
    
    print(tol, status)
    print(objects)

    # save the statistics data dict as json
    with open(stats_raw_path, 'w') as f:
        json.dump(stats, f, indent=4)

    # save the object counts dict as json
    with open(obj_dist_path, 'w') as f:
        json.dump(objects, f, indent=4)
        
    # save the Excel book of statistics data
    wb.save(stats_sheet_path)
