'''
annotate.py

Lin Li (linli.tree@outlook.com) / 21 Sep. 2022 (last modified)

Automation tool for annotating the processed data

Contributions:

The entire code was written by Lin. 


'''

import os, sys, csv, json
import tkinter as tk
import random
from tkinter.messagebox import showinfo
from tkinter import font
from math import sqrt
import datetime as dt

from PIL import Image, ImageTk, ImageDraw, ImageFont
from addict import Dict
from openpyxl import load_workbook
from argparse import ArgumentParser

curr_path = os.path.dirname(os.path.abspath(__file__)) # the directory of the current file
cwp =  os.path.join(curr_path, "..")
# add root path to the system PATH to allow import other user-defined modules
sys.path.append(cwp)

from src.log import  *
from src.log import save as save_log


'''
Arguments and Constants

'''

argparser = ArgumentParser()
argparser.add_argument('--takes', nargs='+', default=None,
                       help="ID of takes to be annotated. Set None to annotate all takes in the 'data' directory. This can be given with a single integer number for one take or a range linked by '-', e.g., '10-12' for takes [000010, 000011, 000012]")
argparser.add_argument('--review', action='store_true', default=False,
                       help="true to review the takes have been already annotated before. By default (False), the annotated takes will not be displayed for annotation again.")
argparser.add_argument('--datapath', type=str, default="Sample_Cases/data",
                       help='data path to all takes')

# display configuration of each stream
# key: stream ID
# values:
#     img_path: path template to the image
#     dev: device ID
#     ts_path: path to the timestamp file
#     position: position in the grid layout
STREAMS = Dict({
    'rgbd0_rgb' : {'img_path' : 'rgbd0/left_{:04d}.png',
                   'dev' : 'rgbd0',
                   'ts_path' : 'rgbd0_ts.csv',
                   'position' : (0, 0)},
    'rgbd0_depth' : {'img_path' : 'rgbd0/depth_{:04d}.png',
                     'dev' : 'rgbd0',
                     'ts_path' : 'rgbd0_ts.csv',
                     'position' : (0, 1)},
    'rgbd1_rgb' : {'img_path' : 'rgbd1/left_{:04d}.png',
                   'dev' : 'rgbd1',
                   'ts_path' : 'rgbd1_ts.csv',
                   'position' : (1, 0)},
    'rgbd1_depth' : {'img_path' : 'rgbd1/depth_{:04d}.png',
                     'dev' : 'rgbd1',
                     'ts_path' : 'rgbd1_ts.csv',
                     'position' : (1, 1)},
    'rgbd2_rgb' : {'img_path' : 'rgbd2/left_{:04d}.png',
                   'dev' : 'rgbd2',
                   'ts_path' : 'rgbd2_ts.csv',
                   'position' : (2, 0)},
    'rgbd2_depth' : {'img_path' : 'rgbd2/depth_{:04d}.png',
                     'dev' : 'rgbd2',
                     'ts_path' : 'rgbd2_ts.csv',
                     'position' : (2, 1)},
    'event' : { 'img_path' : 'event/{:04d}.jpg',
                'dev' : 'event',
                'ts_path' : 'event_frames_ts.csv',
                'position' : (0, 2)},
    'hand' : {'img_path' : 'hand_motion/{}.jpg',
              'dev' : 'left_hand_pose',
              'ts_path' : None,
              'position': (1, 2)}
})

# annotatin status mapping
STATUS = {
    0 : 'unfinished',
    1 : 'finished',
    -1 : 'problematic'
}

# default Timestamp display
T = Dict({'rgbd0':{'frame':0, 'timestamp':'0000000000000000000'}})
# default Hand display
H = 'hand'
# default Horizontal dislay
HOR = 'horizontal'
# default Vertical display
HEG = 'vertical'
# default Position display
POS = '(x, z)'
# default Speed display
SPEED = 0

# options of Hand
HAND = ['left', 'right', 'both']

data = {}


'''
Auxiliary functions

'''

def load_timestamp(path):
    '''
    Load timestamp file

    Returns:
        list: of int timestamps

    Args:
        path (string): path to the timestamp file

    '''
    
    with open(path, 'r') as f:
        # ignore the first row since it is the header
        return [int(ts) for ts in f.readlines()[1:]]

def load_anno(path):
    '''
    Load annotation data

    Returns:
        Dict: annotation data

    Args:
        path (string): path to the annotation data file

    '''
    
    with open(path, 'r') as f:
        # load dict from json file
        return Dict(json.loads(f.read()))

def init_anno(path, take):
    '''
    Initialize annotation data (empty)

    Returns:
        Dict: annotation data dict with default values

    Args:
        path (string): path to the annotation data file
        take (dict): take data

    '''

    if os.path.isfile(path): # annotation data file exists
        anno = load_anno(path) # load annotation data from the file
    else: # not exist
        anno = Dict({'status':0}) # initialize with a new Dict

        # fill anno dict with the key-values from the take data
        for k, v in take.items():
            # ignore the row information
            if 'row' in k: continue
            anno[k] = v

        # initialize hand status, position, moments with the default value
        anno.throw.hand = H
        anno.throw.hand_vertical_thrower = HEG
        anno.throw.hand_vertical_catcher = HEG
        anno.throw.hand_horizontal_thrower = HOR
        anno.throw.hand_horizontal_catcher = HOR
        anno.throw.position_thrower = POS
        anno.throw.position_catcher = POS
        anno.throw.object_flying_speed = SPEED
        
        anno.catch.hand = H
        anno.catch.hand_vertical = HEG
        anno.catch.hand_horizontal = HOR
        anno.catch.position = POS
        
        anno.throw.time_point = Dict(T)
        anno.catch.time_point_touch = Dict(T)
        anno.catch.time_point_stable = Dict(T)
        
    return anno

def init_data(take, anno_path, proc_path):
    '''
    Initialize annotation data

    Returns:
        dict: annotation data
        dict: alignment timestamps
        int: frame index
        dict: optitarck data

    Args:
        take (string): take ID
        anno_path (string): path to the annotation file
        proc_path (string): path to the processed directory

    '''

    # initialize annotation data
    anno = init_anno(anno_path, take)

    tss = {} # timestamps collection for all streams from the timestamp files
    for stream, cfg in STREAMS.items():
        dev = cfg.dev # device ID
        # skip the existing device and hand
        if dev in tss or stream=='hand': continue
        # load timestamp
        ts = load_timestamp(os.path.join(proc_path, cfg.ts_path))
        tss[dev] = ts

    frame = 0 # initial frame index

    align_path = os.path.join(proc_path, 'alignment.json') # path to the alignment file
    with open(align_path, 'r') as f: align = json.loads(f.read()) # load alignment data
    # aligned timestamps for each stream
    # key: device ID
    # value: dict of frames
    #     key: frame index
    #     value: dict
    #         key: frame, timestamp
    aligned = {} 
    for fnum, dev_ts in align.items(): # iterate every frame in alignment data
        for dev, ts in dev_ts.items(): # iterate every device in that frame
            # initialize device frame data if device not exists in aligned
            if dev not in aligned: aligned[dev] = [Dict() for _ in align.keys()]
            fnum = int(fnum) # string frame number to int

            # get frame index in timestamp file if device has the timestamp file
            # otherwise, use the frame index from alignment data
            aligned[dev][fnum].frame = tss[dev].index(ts) if dev in tss else fnum
            aligned[dev][fnum].timestamp = ts

    # optitack data
    # key: object ID
    # value: dict:
    #     key: timestamp
    #     value: optitrack data list
    opti = {}
    # only retrieve the optitrack data for the subjects' head motion
    # to get the location of two subjects and compute the flying speed of the thrown-away object
    for oid in ['sub1_head_motion', 'sub2_head_motion']:
        opti_path = os.path.join(proc_path, oid+'.csv') # path to optitrack data file
        with open(opti_path, 'r') as f:
            opti[oid] = {} # new sub dict for a object ID
            for line in f.readlines()[1:]: # iterate every row from 2nd
                line = line.split(',') # a list of separated data
                # timestamp as key, the rest data as value
                opti[oid][int(line[0])] = line[1:] 
                
    return anno, aligned, frame, opti


def save_anno(data, path):
    '''
    Save annotation data to the file

    Returns:
        None

    Args:
        data (Dict): annotation data
        path (string): path to save the annotation data

    '''
    
    with open(path, 'w') as f:
        # save dict in json format
        json.dump(data, f, indent=4) # 4 indent to improve legibility

def is_default_t(t):
    '''
    Check if the moment value is default

    Returns:
        bool: true for is default value

    Args:
        t (Dict): moment data dict
    
    '''

    # True if the timestamp value equals to the default timestamp value
    return t.rgbd0.timestamp == T.rgbd0.timestamp

def loop_select(v, options):
    '''
    Circulate among the given options

    Returns:
        auto: the selected value

    Args:
        v (auto): the current value
        options (list): all available options

    '''
    
    if v not in options: # the current value is the default one
        return options[0]

    idx = options.index(v) # the index of current value in the optionss
    idx += 1 # increase the index by 1
    # set index to 0 if it exceeds the maximum number of options
    idx = idx if idx < len(options) else 0
    return options[idx]

def is_anno_finished(anno):
    '''
    Check if the annotation has been finished i.e. all data field (value) are not default

    Returns:
        bool: true for all data field non-default
    
    Args:
        anno (Dict): annotation data

    '''
    
    # list of bool; each holds if data field is default
    finished = []
    
    # check moments
    finished.append(not is_default_t(anno.throw.time_point))
    finished.append(not is_default_t(anno.catch.time_point_touch))
    finished.append(not is_default_t(anno.catch.time_point_stable))

    # check hand status, position, etc.
    finished.append(anno.throw.hand != H)
    finished.append(anno.throw.position_thrower != POS)
    finished.append(anno.throw.position_catcher != POS)
    finished.append(anno.throw.hand_vertical_thrower != HEG)
    finished.append(anno.throw.hand_vertical_catcher != HEG)
    finished.append(anno.throw.hand_horizontal_thrower != HOR)
    finished.append(anno.throw.hand_horizontal_catcher != HOR)
    
    finished.append(anno.catch.hand != H)
    finished.append(anno.catch.position != POS)
    finished.append(anno.catch.hand_vertical != HEG)
    finished.append(anno.catch.hand_horizontal != HOR)

    finished.append(anno.throw.object_flying_speed != 0)

    # return True if all checks return True
    return all(finished)


'''
Interface UI

'''

class ImageBlock(tk.Frame):
    '''
    Frame image displayer

    '''

    # default system font in Ubuntu system
    FONT = ImageFont.truetype("/usr/share/fonts/truetype/freefont/FreeMonoBold.ttf", 40)
    
    def __init__(self, master, dev, img_path, frame, tss, width=200, height=200):
        '''
        Initialization of the displayer

        Returns:
            None

        Args:
            master (tkinter): the parent tkinter container of the current one
            dev (string): device ID
            img_path (string): path tempalte to the frame images
            frame (int): current frame index
            tss (list): sequence of timestamps of the device 'dev'
            width (int): initial width of the displayer
            height (int): initial height of the displayer

        '''

        # initialize with the parent class's initializer
        super(ImageBlock, self).__init__(master,
                                         relief=tk.RAISED, # set the border with raised animation
                                         borderwidth=1,
                                         width=width,
                                         height=height)
        # the backbone displayer is a tk.Label
        self.display = tk.Label(master=self,
                                bg='red', # background color
                                width=width,
                                height=height)
        # set layout to be automatically adjusted according to the parent container
        self.display.pack(fill=tk.BOTH, expand=True)

        self.dev = dev

        # bind <Configure> event (parent resize) to self.resize to customize the resize effect
        self.bind("<Configure>", self.resize)
        # further initialize by reset()
        self.reset(img_path, frame, tss)
        
    def reset(self, img_path, frame, tss):
        '''
        Reset the content

        Returns:
            None
        
        Args:
            img_path (string): path tempalte to the frame images
            frame (int): current frame index
            tss (list): sequence of timestamps of the device 'dev'
        
        '''
        
        self.img_path = img_path
        self.frame = frame
        self.nframes = len(tss) # total number of frames
        self.tss = tss

        # placehoder image collections
        self.imgs = [None for _ in range(self.nframes)]
        # placehoder photo (image) collections
        self.photos = [None for _ in range(self.nframes)]

        # display the current frame
        self.display_frame(frame)

    def resize(self, event):
        '''
        Resize the current UI. This is called when the parent UI changes its size.

        Returns:
            None

        Args:
            event (tkinter.Event): event of size change from the parent container
        
        '''

        # the new size to be resized to
        size = (event.width, event.height)
        # resize the current frame image
        resized = self.load_image(self.frame).resize(size, Image.Resampling.LANCZOS)
        # generate new photo with the resized image
        photo = ImageTk.PhotoImage(resized)
        # display the resized iamge
        self.display.configure(image=photo)
        self.display.image = photo
        # store the photo for reuse
        self.photos[self.frame] = photo

    def display_frame(self, frame):
        '''
        Display the image of a frame
        
        Returns:
            None
       
        Args:
            frame (int): the frame index whose image to display

        '''

        img = self.load_image(frame) # the frame image to display
        # the size of the current window
        size = (self.display.winfo_width(), self.display.winfo_height())
        # the photo of the frame
        photo = self.photos[frame]
        # create new photo if the existing one is None or the size doesn't match
        if photo is None or size != (photo.width(), photo.height()):
            # resize the image to the window size
            img = img.resize(size, Image.Resampling.LANCZOS)
            # generate new photo with resized image
            photo = ImageTk.PhotoImage(img)
            # store photo for reuse
            self.photos[frame] = photo

        # display the image
        self.display.configure(image=photo)
        self.display.image = photo
        
    def load_image(self, frame):
        '''
        Load the image of a frame

        Returns:
            None

        Args:
            frame (int): the frame index

        '''
        
        if self.imgs[frame] is None: # the frame image doesn't exist in the collection
            img_id = self.tss[frame]['frame'] # the frame index in the timestamp file as image ID
            ts = self.tss[frame]['timestamp'] # the timestamp of the frame
            img_path = self.img_path.format(img_id) # path to the frame image
            img = Image.open(img_path) # load image
            draw = ImageDraw.Draw(img) # drawing object
            if ts is None: # no timestamp available
                ts = ''
            else:
                # human-readable format of the timestamp
                date = dt.datetime.fromtimestamp(int(ts/1e9))
                date_str = date.strftime('%m.%d %H:%M:%S')
                ts = date_str + ':{:03d}'.format(int(ts%1e9/1e6))
    
            info = '{}/{} {}'.format(frame, self.nframes, ts) # timestamp information to display
            # draw the timestamp infomation on the image
            draw.text((10, 10), info, 'red', font=self.FONT)
            # draw the timestamp info using the system default font to work on Windows 
            # draw.text((10, 10), info, 'red')
            # store the image for reuse
            self.imgs[frame] = img
        else:
            # reuse the existing image if have
            img = self.imgs[frame]
        return img

    
    def forward(self):
        '''
        Forward a frame

        Returns:
            None

        Args:
            None

        '''

        # increase the frame index by 1
        self.frame = min(self.nframes, self.frame + 1)
        # display the new frame
        self.display_frame(self.frame)
        
    def backward(self):
        '''
        Backward a frame

        Returns:
            None

        Args:
            None

        '''

        # decrease the frame index by 1
        self.frame = max(0, self.frame - 1)
        # display the new frame
        self.display_frame(self.frame)

        
class InfoPanel(tk.Frame):
    '''
    Information panel to display annotation content
    
    '''
    
    def __init__(self, master, anno, width=200, height=200):
        '''
        Initializer

        Returns:
            None

        Args:
        master (tkinter): the parent tkinter container
        anno (dict): the annotation data
        width (int): the width of the window
        height (int): the height of the window

        '''

        # call parent class's initializer
        super(InfoPanel, self).__init__(master,
                                        relief=tk.RAISED, # border animation
                                        borderwidth=1,
                                        width=width,
                                        height=height)
        # stop propagating the layout event to the children UI
        self.pack_propagate(False)
        self.text = tk.StringVar() # string to display
        self.displayer = tk.Label(master=self,
                                  fg='black', # background color black
                                  textvariable=self.text, # bind the display content to the string variable
                                  wraplength=300, # content length
                                  font='Arial 5 bold',
                                  anchor='w', # left alignment
                                  justify=tk.LEFT) # left alignment
        # adapt the size of the label to the container's size automatically 
        self.displayer.pack(fill=tk.BOTH, expand=True, padx=30)

        # content template to display
        self.general_tmp = "status:\t{}\n\n"
        
        self.throw_tmp = "throw:\t{hand:<6}  {position_thrower:<11}  {position_catcher:<11}  {object_flying_speed}ms\n\n"
        self.throw_tmp += "\t{hand_vertical_thrower:<10}   {hand_horizontal_thrower}"
        self.throw_tmp += "\t{hand_vertical_catcher:<10}   {hand_horizontal_catcher}\n\n"
        self.throw_tmp += "\t{time_point.rgbd0.frame:04d}  {time_point.rgbd0.timestamp}\n\n"
        
        self.catch_tmp = "catch:\t{hand:<6}  {position:<11}  {hand_vertical:<10}  {hand_horizontal}\n\n"
        self.catch_tmp += " \t{time_point_touch.rgbd0.frame:04d}   {time_point_touch.rgbd0.timestamp}\n\n"
        self.catch_tmp += " \t{time_point_stable.rgbd0.frame:04d}   {time_point_stable.rgbd0.timestamp}"

        self.info = anno # annotation data
        self.refresh() # refresh content in UI
        
        # bind the parent's resize event to self.resize
        self.bind("<Configure>", self.resize) 

    def resize(self, event):
        '''
        Resize the current UI. This is called when the parent UI changes its size.

        Returns:
            None

        Args:
            event (tkinter.Event): event of size change from the parent container
        
        '''
        
        size = (event.width, event.height) # new size

        # adapt the font size of the text to the new size
        if event.height < 300:
            fsize = 5
        elif event.height < 500:
            fsize = 20
        elif event.height < 800:
            fsize = 30

        # update the display content
        self.displayer.configure(font='Arial {} bold'.format(fsize))
        self.displayer.configure(wraplength=300*fsize/5)

    def format_position(self, pos):
        '''
        Format the position dict to string

        Returns:
            string: position tuple (x, z)

        Args:
            pos (tuple/string): position tuple or string

        '''
        
        if isinstance(pos, str):
            return pos

        return '({}, {})'.format(pos.x, pos.z)
        
    def refresh(self):
        '''
        Refresh the display of the content 
        
        Returns:
            None
        
        Args:
            None

        '''
        
        # throw annotation data
        throw = Dict(self.info.throw)

        # thrower's position at moment throw
        throw.position_thrower = self.format_position(throw.position_thrower)
        # catcher's position at moment throw
        throw.position_catcher = self.format_position(throw.position_catcher)

        # catch annotation data
        catch = Dict(self.info.catch)
        # catcher's position at moment catch(touch)
        catch.position = self.format_position(catch.position)

        # generate content to display according to annotation data
        text = self.general_tmp.format(STATUS[anno.status])
        text += self.throw_tmp.format(**throw)
        text += self.catch_tmp.format(**catch)
        
        # update the text variable in the Label
        self.text.set(text)

    def reset(self, anno):
        '''
        Reset the annotation data

        Returns:
            None

        Args:
            anno (dict): annotation data

        '''
        
        self.info = anno
        # refresh UI to display new annotation data
        self.refresh()

        
'''
Main program

'''

if __name__ == '__main__':
    win = tk.Tk() # root window
    # 3 rows x 3 columns layout
    win.columnconfigure([0, 1, 2], weight=1, minsize=200)
    win.rowconfigure([0, 1, 2], weight=1, minsize=200)

    # parse arguments from console
    args = argparser.parse_args()
    
    # lx
    # folders
    data_path = args.datapath
    root_path = os.path.join(data_path, "..") # code depository root directory
    anno_dir = os.path.join(root_path, 'annotations') # directory to save individual annotation
    # create the annotation directory if not exists
    if not os.path.exists(anno_dir):
        os.makedirs(anno_dir)
    
    
    local_takes = [] # takes available in the local machine
    # an annotatable take should have at least all these data (folders)
    frame_dirs = ['event', 'rgbd0', 'rgbd1', 'rgbd2','hand_motion']
    for take in os.listdir(data_path): # iterate every take folder under data_path
        take_path = os.path.join(data_path, take)
        if all([os.path.exists(os.path.join(take_path, 'processed/'+f)) for f in frame_dirs]):
            # add the take id to the list if all specified data folders exit
            local_takes.append(take)

    # only keep specified takes if explicitly specifying 
    if args.takes is not None:
        args.takes = parse_convert_take_ids(args.takes)
        local_takes = [take for take in args.takes if take in local_takes]

    # get ID of takes annotated and stored in the annotation directory
    # this is used to correct the annotation status in the logbook if inconsistent
    annotated_takes = [take[:-5] for take in os.listdir(anno_dir)]
    
    # find all unannotated takes in the log that are also available in local
    log = load_workbook(log_path)
    takes = Dict() # takes ready to be annotated
    for sub in log.sheetnames: # iterate every subject (sheet)
        sheet = log[sub] # get worksheet of the subject
        for row in range(2, sheet.max_row): # iterate every row (from 2) in the sheet
            take_id = value(sheet, 'take_id', row)
            is_finished = finished(sheet, row) # if the take has been already recorded
            is_annotated = annotated(sheet, row)
            local_available = take_id in local_takes

            # Uncomment lines below to corret the logbook by the annotation data files
            # this should be enabled if the logbook is inconsistent to the annotation data
            # this will not change anything if they are consistent
            
            if take_id in annotated_takes:
                # load annotation if it exists in the annotation directory
                with open(os.path.join(anno_dir, take_id+'.json'), 'r') as f:
                    annotation = json.loads(f.read())
                status = annotation['status']
                if status == 1: # annotated, no issue
                    is_annotated = True
                    set_value(sheet, row, annotated=1, verified=1)
                elif status == 0: # annotated before, but not finished
                    is_annotated = False
                    set_value(sheet, row, annotated='', verified='')
                elif status == -1: # annotated, with issue
                    is_annotated = True
                    vstr = '' if failed(sheet, row) else -1
                    set_value(sheet, row, annotated='', verified=vstr)
                        
            if failed(sheet, row) or problematic(sheet, row) or not local_available:
                # skip takes failed, or having issue, or data not available
                continue

            # only annotate takes that is recorded and not annotated before
            # or annotated but in the review mode
            # if is_finished and (not is_annotated or args.review):
            equipped = value(sheet, 'equipped', row) # if the subject wears the sensors
            hand = value(sheet, 'hand', row)
            position = value(sheet, 'position', row)
            action = value(sheet, 'action', row)
            # logging different information for the subject wearing sensors and not
            if equipped:
                takes[take_id].take_id = take_id
                takes[take_id].object = value(sheet, 'object', row)
                takes[take_id].catch_result = value(sheet, 'success', row)
                takes[take_id].row1 = row # row in the sheet
                takes[take_id].sub1_cmd = Dict({'subject_id' : SUBJECTS.index(sub),
                                                'hand' : hand,
                                                'position' : position,
                                                'action' : action})
            else:
                takes[take_id].sub2_cmd = Dict({'subject_id' : SUBJECTS.index(sub),
                                                'hand' : hand,
                                                'position' : position,
                                                'action' : action,
                                                'hand_vertical' : value(sheet, 'height', row)})
                takes[take_id].row2 = row # row in the sheet
                if action == 'throw':
                    takes[take_id].sub2_cmd.throwing_speed = value(sheet, 'speed', row)
                else:
                    takes[take_id].sub2_cmd.hand_horizontal = value(sheet, 'horizon', row)

    # random shuffle takes for annotation
    take_ids = list(takes.keys())
    random.shuffle(take_ids)
    
    if len(take_ids) == 0:
        print("no takes can be annotated, exit.")
        exit()
    
    anno_idx = 0 # index of the take being annotated
    take_id = take_ids[anno_idx]
    take_path = os.path.join(data_path, take_id)
    proc_path = os.path.join(take_path, 'processed') # the processed directory
    
    take = takes[take_id] # take information

    anno_path = os.path.join(anno_dir, '{}.json'.format(take_id)) # path to the annotation file

    # get annotation data dict, alignment, initial frame no., and optitrack data
    anno, aligned, frame, opti = init_data(take, anno_path, proc_path)
    data[take_id] = [anno, aligned, frame, opti]
    
    # take ID as the title of the interface
    win.title(take_id)

    # collection of displayers for each stream
    # key: stream ID
    # value: displayer UI
    displayers = {}
    for stream, cfg in STREAMS.items():
        # generate full image path template
        img_path = os.path.join(proc_path, cfg.img_path)

        # initialize displayer UI
        display = ImageBlock(win, cfg.dev, img_path, frame, aligned[cfg.dev])
        row, column = cfg.position
        # put displayer at the specified position in the window
        display.grid(row=row, column=column, padx=1, pady=1, sticky='nsew')
        displayers[stream] = display

    # initialize information panel (bottom right)
    info_panel = InfoPanel(win, anno)
    info_panel.grid(row=2, column=2, padx=1, pady=1, sticky='nsew')
    
    def switch_frame(event):
        '''
        Forward or backward a frame in all displayers

        Key mapping
            'right-arrow': forward a frame
            'left-arrow': backward a frame

        Returns:
            None

        Args:
            event (tkinter.Event): key pressing event

        '''
        
        global displayers

        # iterate every displayer
        for dev, dis in displayers.items(): 
            if event.keysym == 'Left':
                # backward a frame if <left-arrow> pressed
                dis.backward()
            else:
                # forward a frame if <right-arrow) pressed
                dis.forward()
                
        # update the current frame index by any displayer's current frame number
        data[take_id][2] = displayers[dev].frame
        
    def add_t(event):
        '''
        Add a moment, update annotation and display
        
        Key mapping
            'Enter': add a moment

        Returns:
            None

        Args:
            event (tkinter.Event): key pressing event

        '''

        # iterate three pre-defined moments in the annotation data in forward order
        for i, t in enumerate([anno.throw.time_point,
                               anno.catch.time_point_touch,
                               anno.catch.time_point_stable]):
            # skip if the moment has been already set
            if not is_default_t(t): continue
            # set up the first unset moment i.e. with default value 
            # get current frame index from any displayer
            frame = next(iter(displayers.values())).frame
            # store timestamps from alignment in the annotation
            for dev, ts in aligned.items():
                t[dev] = ts[frame]
            break # break iteration since only one moment to assign each time

        # determine thrower and catcher according to the action of subject 1
        if anno.sub1_cmd.action == 'throw':
            thrower = 'sub1_head_motion'
            catcher = 'sub2_head_motion'
        else:
            thrower = 'sub2_head_motion'
            catcher = 'sub1_head_motion'

        # a function that converts position tuple to a dict
        pos = lambda x : Dict({'x':round(float(x[0]), 2),
                               'z':round(float(x[2]), 2)})

        # update the information of throw if the moment to be added is 'throw' index 0
        if i == 0:
            # the current timestamp of the thrower's head motion
            ts_throw = t[thrower].timestamp
            # the position of the thrower at that timestamp
            pos_throw = opti[thrower][ts_throw]

            # the current timestamp of the catcher
            ts_catch = t[catcher].timestamp
            # the position of the catcher at that timestamp
            pos_catch = opti[catcher][ts_catch]

            # update thrower and catcher position in the throw section in annotation
            anno.throw.position_thrower = pos(pos_throw)
            anno.throw.position_catcher = pos(pos_catch)

        # update the information of catch if the moment is 'catch(touch)' index 1
        elif i == 1:
            # the timestamp of the moment throw
            ts_throw = anno.throw.time_point[catcher].timestamp
            # the current timestam i.e. the timestamp of the moment catch(touch)
            ts_catch = t[catcher].timestamp
            # the position of catcher at the moment catch(touch)
            pos_catch = opti[catcher][ts_catch]
            pos_catch = pos(pos_catch)
            # update catcher's position in the catch section in annotation
            anno.catch.position = pos_catch

            # the position of thrower at the moment throw
            pos_throw = anno.throw.position_thrower
            # flying duration in seconds
            flying_t = (ts_catch - ts_throw) / 1e9
            # flying distance = sqrt( (x-x')^2 + (z-z')^2)
            flying_s = sqrt((pos_throw.x-pos_catch.x)**2 + (pos_throw.z-pos_catch.z)**2)
            # flying speed = flying distance / (flying duration+1e-9)
            # 1e-9 is a small constant to ensure numerical stability of division
            anno.throw.object_flying_speed = round(flying_s / (flying_t+1e-9), 2)

        # refresh information panel to reflect the new moment
        info_panel.refresh()
        # save the annotation
        save_anno(anno, anno_path)
        
    def remove_t(event):
        '''
        Remove a moment from annotation and update display

        Key mapping
            'Del': remove the latest moment

        Returns:
            None

        Args:
            event (tkinter.Event): key pressing event

        '''

        # iterate three pre-defined moments in the annotation data in backward order
        for i, (info, k) in enumerate([(anno.catch, 'time_point_stable'),
                                       (anno.catch, 'time_point_touch'),
                                       (anno.throw, 'time_point')]):
            # skip if the moment hasn't been set
            if is_default_t(info[k]): continue
            # remove the first non-default moment by assigning default moment value
            info[k] = Dict(T)
            break # break the iteration since only one moment to remove

        # change the annotation status to 0 (unfinished) if it is not -1 (problematic)
        # it actually only changes status from finished 1 to unfinished 0
        anno.status = anno.status if anno.status == -1 else 0

        if i == 1: # the moment to be removed is catch(touch)
            # clear the catcher's position at the moment and flying speed by assigning default values
            # they are determined according to the moment catch(touch)
            anno.catch.position = POS
            anno.throw.object_flying_speed = SPEED
        elif i == 2: # the moment to be remove is throw
            # clear the thrower's and catcher's positions at the moment
            # they are determined according to the moment throw
            anno.throw.position_thrower = POS
            anno.throw.position_catcher = POS

        # refresh information panel to reflect the change
        info_panel.refresh()
        # save the annotation
        save_anno(anno, anno_path)
        
    def select_anno_value(event):
        '''
        Switch the value of hand status: hand, vertical position, horizontal position

        Key mapping
            'q': thrower's hand at moment throw: left, right, both
            'z': catcher's hand at moment catch(touch): left, right, both
            'a': thrower's hand vertical at moment throw: overhead, overhand, chest, underhand
            's': thrower's hand horizontal at moment throw: left, middle, right
            'd': catcher's hand vertical at moment throw: overhead, overhand, chest, underhand
            'f': catcher's hand horizontal at moment throw: left, middle, right
            'c': catcher's hand vertical at moment catch(touch): overhead, overhand, chest, underhand
            'v': catcher's hand horizontal at moment catch(touch): left, middle, right
        
        Returns:
            None

        Args:
            event (tkinter.Event): key pressing event

        '''
        
        event = event.char # the charatect being pressed
        # mapping the pressed key to the annotation entry and options available to switch
        options = {'q' : (anno.throw, 'hand', HAND),
                   'z' : (anno.catch, 'hand', HAND),
                   'a' : (anno.throw, 'hand_vertical_thrower', HEIGHT),
                   's' : (anno.throw, 'hand_horizontal_thrower', HORIZON),
                   'd' : (anno.throw, 'hand_vertical_catcher', HEIGHT),
                   'f' : (anno.throw, 'hand_horizontal_catcher', HORIZON),
                   'c' : (anno.catch, 'hand_vertical', HEIGHT),
                   'v' : (anno.catch, 'hand_horizontal', HORIZON)}

        # dic: annotation section dict
        # k: key to be modified in that section
        # ops: options availabel to switch among
        dic, k, ops = options[event]
        # switch one value among all options
        dic[k] = loop_select(dic[k], ops)

        # refresh the information panel to reflect the change
        info_panel.refresh()
        # save the annotation
        save_anno(anno, anno_path)
        
    def select_status(event):
        '''
        Switch the annotation status 
        
        Key mapping
            'space': switch between 'finished' and 'unfinished'
            'backspace': switch between 'unfinished' and 'problematic'

        Annotation status:
            0: unfinished
            1: finished
            -1: problematic

        Returns:
            None

        Args:
            event (tkinter.Event): key pressing event

        '''
        
        global anno, info_panel, log, take
        
        status = anno.status # current annotation status
        
        if event.keysym == 'space': # 'space' pressed
            if status == -1: # current status is problematic
                showinfo('Warning!',
                         'can not switch between finished and unfinished until problematic is off')

            # current status is unfinished, but the annotation is actually not finished
            # so can't switch the status to finished
            elif status == 0 and not is_anno_finished(anno):
                showinfo('Warning!',
                         'can not be switched to finished because some annotation is not done.')
            else:
                # switch the status between unfinished and finished
                anno.status = loop_select(status, [0, 1])
                
        elif event.keysym == 'BackSpace': # 'backspace' pressed
            # switch between 0 and -1
            anno.status = 0 if anno.status == -1 else -1

        # refresh the information panel to reflect the change
        info_panel.refresh()
        # update logbook
        status = anno.status # the status after modified
        # set 'verified' to be the same as the status (1:verified or -1:problematic)
        # if the current status is not unfinished
        # otherwise, set 'verified' to be empty
        verified = '' if status == 0 else status
        # set 'annotated' to be 1 (finished) if the current status is 1
        # otherwise, empty
        annotated = status if status == 1 else ''
        # iterate two subjects' log info with the index
        for i, sub in enumerate([anno.sub1_cmd, anno.sub2_cmd]):
            # get the sheet of the subject in the logbook
            sheet = log[SUBJECTS[sub.subject_id]]
            # get the row of the take in the sheet
            row = take['row{}'.format(i+1)]
            # update the columns 'verified' and 'annotated'
            set_value(sheet, row, verified=verified)
            set_value(sheet, row, annotated=annotated)
        # save logbook
        save_log(log, log_path)
        # save annotation
        save_anno(anno, anno_path)
        
    def switch_take(event):
        '''
        Switch to next or last take

        Key mapping
            'up-arrow': last take
            'down-arrow': next take

        Returns:
            None

        Args:
            event (tkinter.Event): key pressing event

        '''
        
        global anno_idx, anno, anno_path, aligned, frame, opti, take_id, take_path, proc_path, take
        if event.keysym == 'Down': # 'down-arrow' pressed
            if anno_idx == len(takes) - 1:
                # remind annotation finished since no more takes
                showinfo('good job!', 'no more takes to annotate =^.^=')
            else:
                # warning if the current annotation not finished
                if anno.status == 0:
                    showinfo('wait a minute', 'annotation is not done for {}'.format(take_id))
                # increase the annotation index
                anno_idx += 1
        else:
            # warning if the current annotation not finished
            if anno.status == 0:
                showinfo('wait a minute', 'annotation is not done for {}'.format(take_id))
            # decrease the annotation index, but no less than 0
            anno_idx = max(0, anno_idx - 1)
            
        take_id = take_ids[anno_idx] # new take ID
        take = takes[take_id] # new take information
        take_path = os.path.join(data_path, take_id) # path to the take folder
        proc_path = os.path.join(take_path, 'processed') # path to the processed data
        anno_path = os.path.join(anno_dir, '{}.json'.format(take_id)) # path to the annotation file

        # update window title with the new take ID
        win.title(take_id)

        if take_id in data: # the take has been annotated before in this program run
            # get the annotation material from the existing storage
            anno, aligned, frame, opti = data[take_id]
        else: # the take hasn't been annotated before in this program run
            # initialize all annotation material
            anno, aligned, frame, opti = init_data(take, anno_path, proc_path)
            # store in data dict for share
            data[take_id] = [anno, aligned, frame, opti]

        for stream, displayer in displayers.items(): # iterate every displayer
            cfg = STREAMS[stream] # stream configuration
            img_path = os.path.join(proc_path, cfg.img_path) # update the image path template
            # reset the displayer with the new take data
            displayer.reset(img_path, frame, aligned[cfg.dev])
        # reset the information panel with the new annotation
        info_panel.reset(anno)

    # binding key pressing to the function
    win.bind("<Left>", switch_frame)
    win.bind("<Right>", switch_frame)
    win.bind("<Up>", switch_take)
    win.bind("<Down>", switch_take)
    win.bind("q", select_anno_value)
    win.bind("z", select_anno_value)
    win.bind("a", select_anno_value)
    win.bind("s", select_anno_value)
    win.bind("d", select_anno_value)
    win.bind("f", select_anno_value)
    win.bind("c", select_anno_value)
    win.bind("v", select_anno_value)
    win.bind("<Return>", add_t)
    win.bind("<Delete>", remove_t)
    win.bind("<space>", select_status)
    win.bind("<BackSpace>", select_status)
    
    # launch the window
    win.mainloop()
